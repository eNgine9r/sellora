from __future__ import annotations

from io import BytesIO
from pathlib import Path
import re
from uuid import UUID

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models.import_job import ImportJobStatus
from app.schemas.import_center import ImportReportResponse, ImportValidationIssue, ImportValidationReport
from app.services.import_center_service import ImportServiceError, issue, map_row
from app.services.import_durable_service import (
    DurableImportService,
    append_report_issues,
    append_validation_issues,
    pilot_safe_mapping,
)
from app.services.import_source_storage import ImportSourceStorage


SAFE_SIGNED_NUMBER = re.compile(r"^[+-]?\d+(?:[.,]\d+)?$")
SAFE_INTERNATIONAL_PHONE = re.compile(r"^\+\d{7,15}$")


def is_formula_injection_risk(value: object) -> bool:
    if not isinstance(value, str):
        return False
    text = value.strip()
    if not text:
        return False
    if text[0] not in {"=", "+", "-", "@"}:
        return False
    if SAFE_SIGNED_NUMBER.fullmatch(text) or SAFE_INTERNATIONAL_PHONE.fullmatch(text):
        return False
    return True


def formula_injection_issues(rows: list[dict], mapping: dict[str, str]) -> list[ImportValidationIssue]:
    issues: list[ImportValidationIssue] = []
    for row_number, row in enumerate(rows, start=2):
        for field, column in mapping.items():
            if not column:
                continue
            raw_value = row.get(column)
            if is_formula_injection_risk(raw_value):
                issues.append(issue(row_number, "ERROR", field, "Formula-prefixed CSV values are not allowed", raw_value, None))
    return issues


def append_inventory_update_plan(report: ImportReportResponse, update_rows: list[int]) -> ImportReportResponse:
    """Describe absolute inventory SET operations before execute."""
    if not update_rows:
        return report
    warnings: list[ImportValidationIssue] = []
    for row_number in update_rows:
        warning = issue(row_number, "WARNING", "stock_quantity", "Existing inventory will be updated using absolute quantities")
        report.warnings_by_row.setdefault(row_number, []).append(warning)
        warnings.append(warning)
    report.sample_warnings.extend(warnings[: max(0, 10 - len(report.sample_warnings))])
    report.warnings_count += len(warnings)
    report.warning_rows = len(set(report.warnings_by_row))
    report.updated_count = len(update_rows)
    report.created_count = max(report.ready_to_import_rows - report.updated_count, 0)
    report.estimated_entities_to_create = report.created_count
    return report


class PilotSafeImportService(DurableImportService):
    """Durable Import Center service with controlled-pilot input safety."""

    def __init__(self, db: Session, source_storage: ImportSourceStorage | None = None) -> None:
        super().__init__(db, source_storage=source_storage)

    def _validate_upload_content(self, filename: str, content: bytes) -> None:
        super()._validate_upload_content(filename, content)
        if Path(filename).suffix.lower() != ".xlsx":
            return
        workbook = None
        try:
            workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
            if not workbook.sheetnames:
                raise ImportServiceError("Uploaded .xlsx workbook has no sheets")
        except ImportServiceError:
            raise
        except Exception as exc:
            raise ImportServiceError("Uploaded .xlsx file is not a valid workbook") from exc
        finally:
            if workbook is not None:
                workbook.close()

    def _historical_issues(
        self,
        workspace_id: UUID,
        job_id: UUID,
        entity_type: str,
        sheet_name: str,
        safe_mapping: dict[str, str],
    ) -> list[ImportValidationIssue]:
        issues = list(super()._historical_issues(workspace_id, job_id, entity_type, sheet_name, safe_mapping))
        if entity_type != "orders_history":
            return issues

        job = self._job(workspace_id, job_id)
        _columns, rows = self.parser.read_rows(job.file_path, sheet_name)
        value_report = self.validator.validate(
            entity_type,
            safe_mapping,
            rows,
            workspace_id,
            self.lookup,
        )
        seen = {
            (item.row_number, item.severity, item.field, item.message)
            for item in issues
        }
        for item in value_report.issues:
            key = (item.row_number, item.severity, item.field, item.message)
            if key not in seen:
                issues.append(item)
                seen.add(key)
        return issues

    def _mapped_rows(self, workspace_id: UUID, job_id: UUID, sheet_name: str, column_mapping: dict[str, str]) -> list[dict]:
        job = self._job(workspace_id, job_id)
        _columns, rows = self.parser.read_rows(job.file_path, sheet_name)
        return [self.validator.normalized_row(map_row(row, column_mapping)) for row in rows]

    def _formula_issues(self, workspace_id: UUID, job_id: UUID, entity_type: str, sheet_name: str, column_mapping: dict[str, str]) -> list[ImportValidationIssue]:
        safe_mapping = pilot_safe_mapping(entity_type, column_mapping)
        job = self._job(workspace_id, job_id)
        _columns, rows = self.parser.read_rows(job.file_path, sheet_name)
        return formula_injection_issues(rows, safe_mapping)

    def _inventory_update_rows(self, workspace_id: UUID, job_id: UUID, entity_type: str, sheet_name: str, column_mapping: dict[str, str]) -> list[int]:
        if entity_type != "inventory":
            return []
        update_rows: list[int] = []
        for row_number, data in enumerate(self._mapped_rows(workspace_id, job_id, sheet_name, column_mapping), start=2):
            variant = self.lookup.find_variant(workspace_id, sku=data.get("variant_sku"))
            if variant and self.lookup.inventory_by_variant(workspace_id, variant.id):
                update_rows.append(row_number)
        return update_rows

    def validate(self, workspace_id: UUID, job_id: UUID, entity_type: str, sheet_name: str, column_mapping: dict[str, str], actor_user_id: UUID | None, options: dict | None = None) -> ImportValidationReport:
        report = super().validate(workspace_id, job_id, entity_type, sheet_name, column_mapping, actor_user_id, options)
        return append_validation_issues(report, self._formula_issues(workspace_id, job_id, entity_type, sheet_name, column_mapping))

    def dry_run(self, workspace_id: UUID, job_id: UUID, entity_type: str, sheet_name: str, column_mapping: dict[str, str], actor_user_id: UUID | None = None, options: dict | None = None) -> ImportReportResponse:
        report = super().dry_run(workspace_id, job_id, entity_type, sheet_name, column_mapping, actor_user_id, options)
        unsafe = self._formula_issues(workspace_id, job_id, entity_type, sheet_name, column_mapping)
        append_report_issues(report, unsafe)
        append_inventory_update_plan(report, self._inventory_update_rows(workspace_id, job_id, entity_type, sheet_name, column_mapping))
        if unsafe:
            job = self._job(workspace_id, job_id)
            job.status = ImportJobStatus.FAILED.value
            self.db.commit()
        return report

    def execute(self, workspace_id: UUID, job_id: UUID, entity_type: str, sheet_name: str, column_mapping: dict[str, str], mode: str, actor_user_id: UUID | None, dry_run: bool = False, options: dict | None = None):
        if not dry_run and self._formula_issues(workspace_id, job_id, entity_type, sheet_name, column_mapping):
            raise ImportServiceError("Formula-prefixed values are not allowed; correct the source file and run dry-run again")
        return super().execute(workspace_id, job_id, entity_type, sheet_name, column_mapping, mode, actor_user_id, dry_run, options)
