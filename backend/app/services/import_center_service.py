from __future__ import annotations

from collections import defaultdict
from datetime import UTC, date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from re import split, sub
from uuid import UUID

from fastapi import UploadFile
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.config import get_settings
from app.models.ad_campaign import AdCampaign, AdCampaignBudgetType, AdCampaignObjective, AdCampaignPlatform, AdCampaignStatus
from app.models.ad_metric import AdMetric
from app.models.customer import Customer
from app.models.import_job import ImportJob, ImportJobStatus
from app.models.import_job_log import ImportJobLog, ImportJobLogStatus
from app.models.inventory import Inventory
from app.models.order import Order, OrderStatus, PaymentStatus
from app.models.product import Product
from app.models.product_image import ProductImage
from app.models.product_variant import ProductVariant
from app.models.shipment import Shipment, ShipmentCarrier, ShipmentStatus
from app.repositories.audit_log_repository import AuditLogRepository
from app.repositories.import_center_repository import ImportEntityLookupRepository, ImportJobLogRepository, ImportJobRepository
from app.schemas.import_center import ImportReportResponse, ImportValidationIssue, ImportValidationReport, SuggestMappingResponse, YourJewelryPresetResponse
from app.schemas.advertising import AdCampaignCreate, AdMetricCreate
from app.schemas.order import OrderCreate, OrderItemCreate
from app.services.advertising_service import AdCampaignService, AdMetricService, AdvertisingServiceError
from app.services.business_utils import snapshot
from app.services.order_service import OrderService, OrderServiceError

SUPPORTED_ENTITY_TYPES = {"customers", "products", "product_variants", "inventory", "orders", "ad_campaigns", "ad_metrics", "shipments", "product_catalog", "orders_history", "advertising_history"}
YOUR_JEWELRY_SHEETS = ["Замовлення 2022-2025", "Аналітика Реклами 2023-2025", "Наявність на складі годинників", "Main Watchh інфа про товар"]

PRODUCT_CATALOG_PRESET_NAME = "your_jewelry_product_catalog_v1"
PRODUCT_CATALOG_MAPPING = {
    "product_sku": "Родительский артикул",
    "product_sku_fallback": "Артикул для отображения на сайте",
    "product_name": "Название (UA)",
    "product_name_fallback": "Название модификации (UA)",
    "category": "Раздел",
    "description": "Описание товара (UA)",
    "brand": "Бренд",
    "is_active": "Отображать",
    "variant_sku": "Артикул",
    "color": "Цвет",
    "size": "Розмір",
    "selling_price": "Цена",
    "barcode": "Штрихкод",
    "currency": "Валюта",
    "availability": "Наличие",
    "quantity": "Количество",
    "primary_image_url": "Фото",
    "gallery_urls": "Галерея",
}
PRODUCT_CATALOG_COLUMNS = list(dict.fromkeys(PRODUCT_CATALOG_MAPPING.values()))

ORDERS_HISTORY_PRESET_NAME = "your_jewelry_orders_history_v1"
ADVERTISING_HISTORY_PRESET_NAME = "your_jewelry_advertising_history_v1"
ORDERS_HISTORY_MAPPING = {
    "order_number": "Order Number",
    "order_date": "Order Date",
    "customer_name": "Customer Name",
    "customer_phone": "Customer Phone",
    "instagram_username": "Instagram",
    "variant_sku": "Variant SKU",
    "product_name": "Product Name",
    "quantity": "Quantity",
    "unit_price": "Unit Price",
    "unit_cost": "Unit Cost",
    "order_total": "Total",
    "ad_cost": "Ad Cost",
    "shipping_cost": "Shipping Cost",
    "cod_fee": "COD Fee",
    "other_cost": "Other Cost",
    "payment_status": "Payment Status",
    "order_status": "Order Status",
    "tracking_number": "Tracking Number",
    "carrier": "Carrier",
    "city": "City",
    "warehouse": "Warehouse",
    "notes": "Notes",
}
ADVERTISING_HISTORY_MAPPING = {
    "campaign_name": "Campaign Name",
    "platform": "Platform",
    "metric_date": "Date",
    "spend": "Spend",
    "impressions": "Impressions",
    "reach": "Reach",
    "clicks": "Clicks",
    "messages": "Messages",
    "leads": "Leads",
    "orders": "Orders",
    "revenue": "Revenue",
    "net_profit": "Net Profit",
    "notes": "Notes",
}
FIELD_ALIASES: dict[str, dict[str, list[str]]] = {
    "customers": {
        "name": ["Імʼя", "Ім'я", "Клієнт", "Customer", "Name", "name"],
        "phone": ["Телефон", "Phone", "customer_phone", "phone"],
        "instagram_username": ["Instagram", "Інстаграм", "instagram_username"],
        "city": ["Місто", "City", "city"],
        "region": ["Область", "Region", "region"],
    },
    "products": {
        "name": ["Назва", "Товар", "Product", "Product Name", "name", "Название (UA)", "Название модификации (UA)"],
        "sku": ["Артикул", "SKU", "product_sku", "sku", "Родительский артикул", "Артикул для отображения на сайте"],
        "description": ["Опис", "Description", "description", "Описание товара (UA)"],
        "category": ["Раздел", "Category", "category"],
        "brand": ["Бренд", "Brand", "brand"],
        "is_active": ["Отображать", "Active", "is_active"],
    },
    "product_catalog": {field: [column] for field, column in PRODUCT_CATALOG_MAPPING.items()},
    "product_variants": {
        "product_name": ["Товар", "Назва", "Product", "product_name"],
        "product_sku": ["Артикул товару", "product_sku"],
        "variant_sku": ["Модель годинника", "Модель", "variant_sku", "SKU", "Артикул"],
        "color": ["Колір", "Color", "color"],
        "size": ["Розмір", "Size", "size"],
        "selling_price": ["Ціна", "Selling Price", "selling_price"],
    },
    "inventory": {
        "variant_sku": ["Модель годинника", "Модель", "variant_sku", "SKU", "Артикул"],
        "stock_quantity": ["Залишок", "stock_quantity", "Кількість"],
        "reserved_quantity": ["Зарезервовано", "reserved_quantity"],
        "incoming_quantity": ["У дорозі", "incoming_quantity"],
        "minimum_quantity": ["Мінімальний залишок", "minimum_quantity"],
    },
    "ad_campaigns": {
        "name": ["Кампанія", "Campaign", "Campaign Name", "campaign_name", "name"],
        "platform": ["Платформа", "Platform", "platform"],
        "objective": ["Ціль", "Objective", "objective"],
        "budget_type": ["Тип бюджету", "Budget Type", "budget_type"],
        "daily_budget": ["Денний бюджет", "Daily Budget", "daily_budget"],
        "total_budget": ["Загальний бюджет", "Total Budget", "total_budget"],
        "start_date": ["Дата старту", "Start Date", "start_date"],
        "end_date": ["Дата завершення", "End Date", "end_date"],
        "notes": ["Нотатки", "Notes", "notes"],
    },
    "ad_metrics": {
        "campaign_name": ["Кампанія", "Campaign", "campaign_name"],
        "campaign_id": ["campaign_id"],
        "metric_date": ["Дата", "Date", "metric_date"],
        "spend": ["Витрати на рекламу", "Реклама", "Spend", "ad_spend", "spend"],
        "impressions": ["Покази", "Impressions", "impressions"],
        "reach": ["Охоплення", "Reach", "reach"],
        "clicks": ["Кліки", "Clicks", "clicks"],
        "messages": ["Повідомлення", "Звернення", "Direct", "Messages", "messages"],
        "leads": ["Ліди", "Leads", "leads"],
        "orders": ["Замовлення", "Orders", "orders"],
        "revenue": ["Виручка", "Дохід", "Revenue", "revenue"],
        "net_profit": ["Прибуток", "Profit", "net_profit"],
    },

    "shipments": {
        "order_number": ["Номер замовлення", "Order Number", "order_number"],
        "tracking_number": ["ТТН", "Номер ТТН", "Tracking Number", "tracking_number"],
        "carrier": ["Перевізник", "Carrier", "carrier"],
        "status": ["Статус доставки", "Shipment Status", "status"],
        "recipient_name": ["Отримувач", "Recipient Name", "recipient_name"],
        "recipient_phone": ["Телефон", "Recipient Phone", "recipient_phone"],
        "city": ["Місто", "City", "city"],
        "warehouse": ["Відділення", "Warehouse", "warehouse"],
        "shipping_cost": ["Вартість доставки", "Shipping Cost", "shipping_cost"],
        "cod_amount": ["Накладений платіж", "COD Amount", "cod_amount"],
        "declared_value": ["Оголошена вартість", "Declared Value", "declared_value"],
        "notes": ["Нотатки", "Notes", "notes"],
    },
    "orders": {
        "customer_name": ["Клієнт", "Customer", "customer_name"],
        "customer_phone": ["Телефон", "Phone", "customer_phone"],
        "instagram_username": ["Instagram", "instagram_username"],
        "created_at": ["Дата замовлення", "Дата", "Order Date", "created_at"],
        "order_date": ["Дата замовлення", "Дата", "Order Date", "order_date"],
        "revenue": ["Сума замовлення", "Сума", "Revenue", "revenue"],
        "order_total": ["Сума замовлення", "Сума", "Order Total", "order_total"],
        "net_profit": ["Прибуток", "Profit", "net_profit"],
        "ad_cost": ["Витрати на рекламу", "Реклама", "ad_cost"],
        "shipping_cost": ["Доставка", "Shipping", "shipping_cost"],
        "cod_fee": ["Накладений платіж", "cod_fee"],
        "other_cost": ["Інші витрати", "other_cost"],
        "city": ["Місто", "City", "city"],
        "region": ["Область", "Region", "region"],
        "product_variant_sku": ["Модель годинника", "Модель", "variant_sku"],
        "quantity": ["Кількість", "Quantity", "quantity"],
    },
    "orders_history": {
        "order_number": ["Номер замовлення", "Замовлення", "Order Number", "order_number", "ID замовлення", "Номер заказа"],
        "order_date": ["Дата", "Дата замовлення", "Created At", "Order Date", "Дата заказа"],
        "customer_name": ["Клієнт", "Покупець", "Customer", "Customer Name", "Ім'я", "Имя", "Name"],
        "customer_phone": ["Телефон", "Phone", "Customer Phone", "Номер телефону"],
        "instagram_username": ["Instagram", "Instagram username", "Інстаграм", "Ник Instagram", "Instagram Nick"],
        "variant_sku": ["Артикул", "SKU", "Variant SKU", "product_variant_sku", "Код товару"],
        "product_name": ["Товар", "Назва товару", "Product", "Product Name", "Название товара"],
        "quantity": ["Кількість", "Количество", "Quantity", "Qty"],
        "unit_price": ["Ціна", "Цена", "Unit Price", "Price", "Ціна продажу"],
        "unit_cost": ["Собівартість", "Закупка", "Unit Cost", "Product Cost", "Cost"],
        "order_total": ["Сума", "Total", "Revenue", "Доход", "Виторг", "Сума замовлення"],
        "ad_cost": ["Реклама", "Вартість реклами", "Ad Cost", "Advertising Cost", "CPA cost"],
        "shipping_cost": ["Доставка", "Shipping", "Shipping Cost", "Вартість доставки"],
        "cod_fee": ["Накладений платіж", "COD", "COD Fee", "Комісія НП", "Комісія післяплати"],
        "other_cost": ["Інші витрати", "Other", "Other Cost", "Додаткові витрати"],
        "payment_status": ["Оплата", "Статус оплати", "Payment Status", "Payment"],
        "order_status": ["Статус", "Статус замовлення", "Order Status", "Status"],
        "tracking_number": ["ТТН", "Номер ТТН", "Tracking Number", "tracking_number"],
        "carrier": ["Перевізник", "Carrier", "Delivery Service", "Служба доставки"],
        "city": ["Місто", "City", "Город"],
        "warehouse": ["Відділення", "Warehouse", "Отделение"],
        "notes": ["Коментар", "Notes", "Comment", "Примітки"],
    },
    "advertising_history": {
        "campaign_name": ["Кампанія", "Назва кампанії", "Campaign", "Campaign Name", "Ad Campaign"],
        "platform": ["Платформа", "Platform", "Джерело реклами"],
        "metric_date": ["Дата", "Date", "Metric Date", "День"],
        "spend": ["Витрати", "Spend", "Ad Spend", "Реклама", "Вартість реклами"],
        "impressions": ["Покази", "Impressions"],
        "reach": ["Охоплення", "Reach"],
        "clicks": ["Кліки", "Clicks"],
        "messages": ["Повідомлення", "Messages", "Direct", "Діалоги"],
        "leads": ["Ліди", "Leads"],
        "orders": ["Замовлення", "Orders", "Purchases"],
        "revenue": ["Дохід", "Revenue", "Sales", "Виторг"],
        "net_profit": ["Чистий прибуток", "Net Profit", "Profit"],
        "notes": ["Notes", "Коментар", "Примітки"],
    },
}
YOUR_JEWELRY_EXCEL_V1 = {
    "name": "your_jewelry_excel_v1",
    "sheets": YOUR_JEWELRY_SHEETS,
    "mappings": {
        "customers": {"name": "Імʼя", "phone": "Телефон", "instagram_username": "Instagram", "city": "Місто", "region": "Область"},
        "products": {"name": "Назва", "sku": "Артикул", "description": "Опис"},
        "inventory": {"variant_sku": "Артикул", "stock_quantity": "Кількість", "minimum_quantity": "Мінімальний залишок"},
        "orders": {"customer_name": "Клієнт", "customer_phone": "Телефон", "revenue": "Сума", "created_at": "Дата", "city": "Місто", "region": "Область"},
        "ad_campaigns": {"name": "Кампанія", "platform": "Платформа", "objective": "Ціль"},
        "ad_metrics": {"campaign_name": "Кампанія", "metric_date": "Дата", "spend": "Витрати на рекламу", "impressions": "Покази", "reach": "Охоплення", "clicks": "Кліки", "messages": "Повідомлення", "leads": "Ліди", "orders": "Замовлення", "revenue": "Виручка", "net_profit": "Прибуток"},
        "shipments": {"order_number": "Номер замовлення", "tracking_number": "ТТН", "carrier": "Перевізник", "status": "Статус доставки", "city": "Місто", "warehouse": "Відділення", "recipient_phone": "Телефон", "recipient_name": "Отримувач"},
        "orders_history": ORDERS_HISTORY_MAPPING,
        "advertising_history": ADVERTISING_HISTORY_MAPPING,
    },
}


class ImportServiceError(ValueError):
    pass


class ExcelValueNormalizer:
    empty_tokens = {"", "-", "—", "–", "n/a", "none", "null"}

    def text(self, value) -> str | None:
        if value is None:
            return None
        if isinstance(value, str):
            cleaned = " ".join(value.strip().split())
            return None if cleaned.lower() in self.empty_tokens else cleaned
        return str(value).strip()

    def number(self, value) -> Decimal | None:
        if value is None:
            return None
        if isinstance(value, bool):
            return Decimal(int(value))
        if isinstance(value, int | float | Decimal):
            return Decimal(str(value))
        text = self.text(value)
        if text is None:
            return None
        text = text.replace("%", "")
        text = sub(r"[^0-9,\.\-]", "", text.replace(" ", ""))
        if text.count(",") == 1 and text.count(".") == 0:
            text = text.replace(",", ".")
        elif text.count(",") > 0 and text.count(".") > 0:
            text = text.replace(",", "")
        try:
            return Decimal(text)
        except (InvalidOperation, ValueError):
            return None

    def date(self, value) -> datetime | None:
        if value in (None, ""):
            return None
        if isinstance(value, datetime):
            return value if value.tzinfo else value.replace(tzinfo=UTC)
        if isinstance(value, date):
            return datetime(value.year, value.month, value.day, tzinfo=UTC)
        if isinstance(value, int | float) and value > 0:
            from datetime import timedelta
            # Excel serial dates use 1899-12-30 as the common epoch.
            return datetime(1899, 12, 30, tzinfo=UTC) + timedelta(days=float(value))
        text = self.text(value)
        if text is None:
            return None
        formats = ["%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d", "%Y-%m-%dT%H:%M:%S", "%d.%m.%Y %H:%M:%S"]
        for fmt in formats:
            try:
                parsed = datetime.strptime(text, fmt)
                return parsed.replace(tzinfo=UTC)
            except ValueError:
                continue
        try:
            parsed = datetime.fromisoformat(text)
            return parsed if parsed.tzinfo else parsed.replace(tzinfo=UTC)
        except ValueError:
            return None

    def boolean(self, value) -> bool | None:
        text = self.text(value)
        if text is None:
            return None
        if text.lower() in {"true", "yes", "так", "1"}:
            return True
        if text.lower() in {"false", "no", "ні", "0"}:
            return False
        return None

    def percent(self, value) -> Decimal | None:
        number = self.number(value)
        return None if number is None else number / Decimal("100")


class ExcelParserService:
    def __init__(self) -> None:
        self.normalizer = ExcelValueNormalizer()

    def list_sheets(self, file_path: str) -> list[str]:
        from openpyxl import load_workbook
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        try:
            return list(workbook.sheetnames)
        finally:
            workbook.close()

    def preview(self, file_path: str, sheet_name: str, limit: int = 20) -> tuple[list[str], list[dict]]:
        return self.read_rows(file_path, sheet_name, limit)

    def read_rows(self, file_path: str, sheet_name: str, limit: int | None = None) -> tuple[list[str], list[dict]]:
        from openpyxl import load_workbook
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        try:
            if sheet_name not in workbook.sheetnames:
                raise ImportServiceError("Sheet not found")
            sheet = workbook[sheet_name]
            iterator = sheet.iter_rows(values_only=True)
            header = next(iterator, None)
            if not header:
                return [], []
            columns = [self.normalizer.text(value) or f"column_{index + 1}" for index, value in enumerate(header)]
            rows: list[dict] = []
            for row in iterator:
                if limit is not None and len(rows) >= limit:
                    break
                if row is None or all(self.normalizer.text(value) is None for value in row):
                    continue
                rows.append({columns[index]: self._clean_cell(value) for index, value in enumerate(row[: len(columns)])})
            return columns, rows
        finally:
            workbook.close()

    def _clean_cell(self, value):
        if isinstance(value, datetime):
            return value.isoformat()
        return self.normalizer.text(value) if isinstance(value, str) else value


class MappingSuggestionService:
    def suggest(self, columns: list[str], entity_type: str) -> SuggestMappingResponse:
        aliases = FIELD_ALIASES.get(entity_type, {})
        suggested: dict[str, str] = {}
        confidence: dict[str, float] = {}
        normalized_columns = {normalize_name(column): column for column in columns}
        for field, field_aliases in aliases.items():
            best_column = None
            best_score = 0.0
            for alias in field_aliases + [field]:
                normalized_alias = normalize_name(alias)
                for normalized_column, original_column in normalized_columns.items():
                    score = 1.0 if normalized_column == normalized_alias else (0.75 if normalized_alias in normalized_column or normalized_column in normalized_alias else 0.0)
                    if score > best_score:
                        best_score = score
                        best_column = original_column
            if best_column and best_score >= 0.75:
                suggested[field] = best_column
                confidence[field] = best_score
        mapped_columns = set(suggested.values())
        required_missing = [", ".join(group) for group in MappingValidationService.required_groups.get(entity_type, []) if not any(field in suggested for field in group)]
        return SuggestMappingResponse(suggested_mapping=suggested, confidence=confidence, unmapped_columns=[column for column in columns if column not in mapped_columns], required_fields_missing=required_missing)

    def product_catalog_preset(self) -> dict:
        return {
            "name": PRODUCT_CATALOG_PRESET_NAME,
            "sheets": ["Product catalog"],
            "mappings": {"product_catalog": PRODUCT_CATALOG_MAPPING},
        }

    def orders_history_preset(self) -> dict:
        return {"name": ORDERS_HISTORY_PRESET_NAME, "sheets": ["Orders history"], "mappings": {"orders_history": ORDERS_HISTORY_MAPPING}}

    def advertising_history_preset(self) -> dict:
        return {"name": ADVERTISING_HISTORY_PRESET_NAME, "sheets": ["Advertising history"], "mappings": {"advertising_history": ADVERTISING_HISTORY_MAPPING}}

    def your_jewelry_preset(self) -> YourJewelryPresetResponse:
        return YourJewelryPresetResponse(
            preset_name="your_jewelry_excel_v1",
            supported_sheets=YOUR_JEWELRY_SHEETS,
            suggested_entity_type_per_sheet={
                "Замовлення 2022-2025": "orders",
                "Аналітика Реклами 2023-2025": "ad_metrics",
                "Наявність на складі годинників": "inventory",
                "Main Watchh інфа про товар": "products",
            },
            suggested_column_mapping_per_sheet={sheet: FIELD_ALIASES for sheet in YOUR_JEWELRY_SHEETS},
            notes=["Preset contains sheet and column aliases only.", "Preview and confirm mappings before import."],
        )


class MappingValidationService:
    required_groups = {
        "customers": [("name", "phone", "instagram_username")],
        "products": [("name",)],
        "product_variants": [("product_name", "product_sku"), ("variant_sku", "color", "size")],
        "inventory": [("variant_sku",), ("stock_quantity",)],
        "orders": [("customer_name", "customer_phone", "instagram_username"), ("revenue", "order_total"), ("created_at", "order_date")],
        "ad_campaigns": [("name",)],
        "ad_metrics": [("campaign_name", "campaign_id"), ("metric_date",)],
        "shipments": [("order_number",)],
        "orders_history": [("variant_sku",), ("quantity",), ("unit_price", "order_total")],
        "advertising_history": [("campaign_name",), ("metric_date",)],
    }
    numeric_fields = {"stock_quantity", "reserved_quantity", "incoming_quantity", "minimum_quantity", "purchase_price", "shipping_cost", "selling_price", "weight", "quantity", "ad_cost", "cod_fee", "other_cost", "net_profit", "revenue", "order_total", "unit_price", "unit_cost", "daily_budget", "total_budget", "spend", "impressions", "reach", "clicks", "messages", "leads", "orders", "cod_amount", "declared_value"}
    non_negative_fields = {"stock_quantity", "reserved_quantity", "incoming_quantity", "minimum_quantity", "quantity", "daily_budget", "total_budget", "spend", "impressions", "reach", "clicks", "messages", "leads", "orders", "revenue", "order_total", "unit_price", "unit_cost", "shipping_cost", "ad_cost", "cod_fee", "other_cost", "cod_amount", "declared_value"}
    date_fields = {"created_at", "order_date", "metric_date", "start_date", "end_date"}

    def __init__(self) -> None:
        self.normalizer = ExcelValueNormalizer()

    def validate(self, entity_type: str, column_mapping: dict[str, str], rows: list[dict], workspace_id: UUID | None = None, lookup: ImportEntityLookupRepository | None = None) -> ImportValidationReport:
        issues: list[ImportValidationIssue] = []
        if entity_type not in SUPPORTED_ENTITY_TYPES:
            issues.append(issue(None, "ERROR", None, "Unsupported entity_type"))
            return report_from_issues(rows, issues)
        for group in self.required_groups[entity_type]:
            if not any(column_mapping.get(field) for field in group):
                issues.append(issue(None, "ERROR", None, f"Required mapping missing: one of {', '.join(group)}"))
        seen_inventory: set[str] = set()
        seen_ad_metrics: set[tuple[object, str]] = set()
        seen_shipments: set[str] = set()
        for row_number, row in enumerate(rows, start=2):
            mapped = map_row(row, column_mapping)
            normalized = self.normalized_row(mapped)
            issues.extend(self._row_required_issues(entity_type, row_number, normalized, mapped))
            for field in self.numeric_fields.intersection(mapped):
                raw = mapped.get(field)
                if raw in (None, ""):
                    continue
                number = self.normalizer.number(raw)
                if number is None:
                    issues.append(issue(row_number, "ERROR", field, f"{field} must be numeric", raw, None))
                elif field in self.non_negative_fields and number < 0:
                    issues.append(issue(row_number, "ERROR", field, f"{field} cannot be negative", raw, number))
            for field in self.date_fields.intersection(mapped):
                raw = mapped.get(field)
                if raw in (None, ""):
                    continue
                parsed = self.normalizer.date(raw)
                if parsed is None:
                    issues.append(issue(row_number, "ERROR", field, f"{field} must be a valid date", raw, None))
            if entity_type == "inventory" and normalized.get("variant_sku"):
                sku = str(normalized["variant_sku"])
                if sku in seen_inventory:
                    issues.append(issue(row_number, "WARNING", "variant_sku", "Duplicate inventory row for variant SKU", None, sku))
                seen_inventory.add(sku)
            if entity_type == "ad_metrics":
                metric_key = (normalized.get("campaign_id") or normalized.get("campaign_name"), str(normalized.get("metric_date")) if normalized.get("metric_date") else None)
                if all(metric_key) and metric_key in seen_ad_metrics:
                    issues.append(issue(row_number, "WARNING", "metric_date", "Duplicate ad metric row for campaign/date", None, metric_key[1]))
                if all(metric_key):
                    seen_ad_metrics.add(metric_key)
            if entity_type == "shipments":
                tracking = normalized.get("tracking_number")
                if tracking:
                    if str(tracking) in seen_shipments:
                        issues.append(issue(row_number, "WARNING", "tracking_number", "Duplicate shipment tracking number in file", None, tracking))
                    seen_shipments.add(str(tracking))
                if normalized.get("status") and normalized.get("status") != ShipmentStatus.DRAFT.value and not tracking:
                    issues.append(issue(row_number, "ERROR", "tracking_number", "tracking_number is required for non-draft shipments"))
            if lookup and workspace_id:
                issues.extend(self._duplicate_issues(entity_type, row_number, normalized, lookup, workspace_id))
        return report_from_issues(rows, issues)

    def normalized_row(self, mapped: dict) -> dict:
        normalized = {}
        for field, value in mapped.items():
            if field in self.numeric_fields:
                normalized[field] = self.normalizer.number(value)
            elif field in self.date_fields:
                normalized[field] = self.normalizer.date(value)
            else:
                normalized[field] = self.normalizer.text(value)
        return normalized

    def _row_required_issues(self, entity_type: str, row_number: int, normalized: dict, mapped: dict) -> list[ImportValidationIssue]:
        issues: list[ImportValidationIssue] = []
        for group in self.required_groups[entity_type]:
            if not any(normalized.get(field) for field in group):
                issues.append(issue(row_number, "ERROR", "/".join(group), f"Row requires one of {', '.join(group)}"))
        return issues

    def _duplicate_issues(self, entity_type: str, row_number: int, normalized: dict, lookup: ImportEntityLookupRepository, workspace_id: UUID) -> list[ImportValidationIssue]:
        if entity_type == "customers" and lookup.find_customer(workspace_id, normalized.get("phone"), normalized.get("instagram_username")):
            return [issue(row_number, "WARNING", "customer", "Duplicate customer in workspace")]
        if entity_type == "products" and lookup.find_product_by_sku(workspace_id, normalized.get("sku")):
            return [issue(row_number, "WARNING", "sku", "Duplicate product SKU in workspace")]
        if entity_type == "product_variants" and lookup.find_variant(workspace_id, sku=normalized.get("variant_sku")):
            return [issue(row_number, "WARNING", "variant_sku", "Duplicate variant SKU in workspace")]
        if entity_type == "ad_campaigns" and lookup.find_ad_campaign_by_name(workspace_id, normalized.get("name")):
            return [issue(row_number, "WARNING", "name", "Duplicate ad campaign in workspace")]
        if entity_type == "ad_metrics":
            campaign = lookup.find_ad_campaign_by_id(workspace_id, normalized.get("campaign_id")) or lookup.find_ad_campaign_by_name(workspace_id, normalized.get("campaign_name"))
            if campaign and normalized.get("metric_date") and lookup.find_ad_metric_by_campaign_date(workspace_id, campaign.id, normalized.get("metric_date").date()):
                return [issue(row_number, "WARNING", "metric_date", "Duplicate ad metric in workspace")]
        if entity_type == "shipments":
            if lookup.find_shipment_by_tracking(workspace_id, normalized.get("tracking_number")):
                return [issue(row_number, "WARNING", "tracking_number", "Duplicate shipment tracking number in workspace")]
            order = lookup.find_order_by_number(workspace_id, normalized.get("order_number"))
            if order is None:
                return [issue(row_number, "ERROR", "order_number", "Order not found in workspace")]
            if lookup.find_shipment_by_order(workspace_id, order.id):
                return [issue(row_number, "WARNING", "order_number", "Active shipment already exists for order")]
        return []


class ProductCatalogImportService:
    availability_aliases = {
        "в наявності": "IN_STOCK", "в наличии": "IN_STOCK", "in stock": "IN_STOCK",
        "немає в наявності": "OUT_OF_STOCK", "нет в наличии": "OUT_OF_STOCK", "out of stock": "OUT_OF_STOCK",
        "очікується": "EXPECTED", "ожидается": "EXPECTED", "expected": "EXPECTED",
    }
    visibility_aliases = {"да": True, "так": True, "yes": True, "true": True, "1": True, "ні": False, "нет": False, "no": False, "false": False, "0": False}

    def __init__(self, db: Session) -> None:
        self.db = db
        self.lookup = ImportEntityLookupRepository(db)
        self.normalizer = ExcelValueNormalizer()

    def normalize_row(self, row: dict, mapping: dict[str, str]) -> dict:
        raw = map_row(row, mapping)
        product_sku = self.normalizer.text(raw.get("product_sku")) or self.normalizer.text(raw.get("product_sku_fallback"))
        product_name = self.normalizer.text(raw.get("product_name")) or self.normalizer.text(raw.get("product_name_fallback"))
        variant_sku = self.normalizer.text(raw.get("variant_sku"))
        price = self.normalizer.number(raw.get("selling_price"))
        quantity = self.normalizer.number(raw.get("quantity"))
        availability = self._availability(raw.get("availability"))
        stock_quantity = int(quantity) if quantity is not None and quantity > 0 else (1 if availability == "IN_STOCK" else 0)
        incoming_quantity = 1 if quantity in (None, 0) and availability == "EXPECTED" else 0
        return {
            "product_sku": product_sku,
            "product_name": product_name,
            "category": self.normalizer.text(raw.get("category")),
            "description": self.normalizer.text(raw.get("description")),
            "brand": self.normalizer.text(raw.get("brand")),
            "is_active": self._visibility(raw.get("is_active")),
            "variant_sku": variant_sku,
            "color": self.normalizer.text(raw.get("color")),
            "size": self.normalizer.text(raw.get("size")),
            "selling_price": price,
            "barcode": self.normalizer.text(raw.get("barcode")),
            "currency": (self.normalizer.text(raw.get("currency")) or "UAH").upper(),
            "availability": availability,
            "stock_quantity": stock_quantity,
            "incoming_quantity": incoming_quantity,
            "minimum_quantity": 0,
            "image_urls": self._image_urls(raw.get("primary_image_url"), raw.get("gallery_urls")),
            "raw_visibility": raw.get("is_active"),
            "raw_availability": raw.get("availability"),
            "raw_currency": raw.get("currency"),
            "raw_price": raw.get("selling_price"),
        }

    def validate(self, rows: list[dict], mapping: dict[str, str], workspace_id: UUID | None = None) -> ImportValidationReport:
        issues: list[ImportValidationIssue] = []
        seen_products: set[str] = set(); seen_variants: set[str] = set()
        for row_number, row in enumerate(rows, start=2):
            data = self.normalize_row(row, mapping)
            if not data["product_sku"]:
                issues.append(issue(row_number, "ERROR", "product_sku", "missing product sku"))
            if not data["product_name"]:
                issues.append(issue(row_number, "ERROR", "product_name", "missing product name"))
            if not data["variant_sku"]:
                issues.append(issue(row_number, "ERROR", "variant_sku", "missing variant sku"))
            if data["selling_price"] is None or data["selling_price"] < 0:
                issues.append(issue(row_number, "ERROR", "selling_price", "invalid price"))
            if data["availability"] is None:
                issues.append(issue(row_number, "WARNING", "availability", "unknown availability"))
            if self.normalizer.text(data["raw_visibility"]) is None:
                issues.append(issue(row_number, "WARNING", "is_active", "visibility missing; defaulting to true"))
            elif data["is_active"] is None:
                issues.append(issue(row_number, "WARNING", "is_active", "invalid visibility value; defaulting to true"))
            if not self.normalizer.text(data["raw_currency"]):
                issues.append(issue(row_number, "WARNING", "currency", "currency missing; defaulting to UAH"))
            elif data["currency"] != "UAH":
                issues.append(issue(row_number, "WARNING", "currency", "currency is not UAH"))
            if not data["category"]:
                issues.append(issue(row_number, "WARNING", "category", "missing category warning"))
            if not data["image_urls"]:
                issues.append(issue(row_number, "WARNING", "image_url", "missing image warning"))
            for url in data["image_urls"]:
                if not self._valid_url(url):
                    issues.append(issue(row_number, "WARNING", "image_url", "invalid image URL"))
            if data["product_sku"]:
                if data["product_sku"] in seen_products:
                    issues.append(issue(row_number, "WARNING", "product_sku", "Duplicate product"))
                seen_products.add(data["product_sku"])
                if workspace_id and self.lookup.find_product_by_sku(workspace_id, data["product_sku"]):
                    issues.append(issue(row_number, "WARNING", "product_sku", "Duplicate product"))
            if data["variant_sku"]:
                if data["variant_sku"] in seen_variants:
                    issues.append(issue(row_number, "WARNING", "variant_sku", "Duplicate variant"))
                seen_variants.add(data["variant_sku"])
                if workspace_id and self.lookup.find_variant(workspace_id, sku=data["variant_sku"]):
                    issues.append(issue(row_number, "WARNING", "variant_sku", "Duplicate variant"))
        return report_from_issues(rows, issues)

    def import_rows(self, workspace_id: UUID, rows: list[dict], mapping: dict[str, str]) -> list[tuple[int, ImportJobLogStatus, str]]:
        results = []
        for row_number, row in enumerate(rows, start=2):
            validation = self.validate([row], mapping, workspace_id)
            if any(item.severity == "ERROR" for item in validation.issues):
                results.append((row_number, ImportJobLogStatus.FAILED, "; ".join(validation.errors)))
                continue
            data = self.normalize_row(row, mapping)
            product = self.lookup.find_product_by_sku(workspace_id, data["product_sku"])
            if product is None:
                product = Product(workspace_id=workspace_id, sku=data["product_sku"], name=data["product_name"], category=data["category"], brand=data["brand"], description=data["description"], is_active=True if data["is_active"] is None else data["is_active"])
                self.db.add(product); self.db.flush()
            variant = self.lookup.find_variant(workspace_id, sku=data["variant_sku"])
            if variant is not None:
                results.append((row_number, ImportJobLogStatus.SKIPPED, "Duplicate variant skipped"))
                continue
            variant = ProductVariant(workspace_id=workspace_id, product_id=product.id, sku=data["variant_sku"], color=data["color"], size=data["size"], price=data["selling_price"], barcode=data["barcode"], is_active=True if data["is_active"] is None else data["is_active"])
            self.db.add(variant); self.db.flush()
            inventory = self.lookup.inventory_by_variant(workspace_id, variant.id)
            if inventory is None:
                inventory = Inventory(workspace_id=workspace_id, product_variant_id=variant.id)
                self.db.add(inventory)
            inventory.stock_quantity = data["stock_quantity"]; inventory.incoming_quantity = data["incoming_quantity"]; inventory.minimum_quantity = 0
            self.db.flush()
            for sort_order, url in enumerate([url for url in data["image_urls"] if self._valid_url(url)]):
                self.db.add(ProductImage(workspace_id=workspace_id, product_id=product.id, image_url=url, sort_order=sort_order, is_primary=sort_order == 0))
            self.db.flush()
            status = ImportJobLogStatus.WARNING if validation.warnings else ImportJobLogStatus.SUCCESS
            results.append((row_number, status, "Product catalog row imported" if status == ImportJobLogStatus.SUCCESS else "Product catalog row imported with warnings"))
        return results

    def metrics(self, rows: list[dict], mapping: dict[str, str], validation: ImportValidationReport) -> dict[str, int]:
        normalized = [self.normalize_row(row, mapping) for row in rows]
        duplicate_products = sum(1 for item in validation.issues if item.field == "product_sku" and "duplicate" in item.message)
        duplicate_variants = sum(1 for item in validation.issues if item.field == "variant_sku" and "duplicate" in item.message)
        return {
            "products_detected": len({item["product_sku"] for item in normalized if item["product_sku"]}),
            "variants_detected": len([item for item in normalized if item["variant_sku"]]),
            "inventory_rows_detected": len([item for item in normalized if item["variant_sku"]]),
            "images_detected": sum(len([url for url in item["image_urls"] if self._valid_url(url)]) for item in normalized),
            "duplicate_products": duplicate_products,
            "duplicate_variants": duplicate_variants,
        }

    def _availability(self, value) -> str | None:
        text = self.normalizer.text(value)
        return self.availability_aliases.get(text.lower()) if text else None

    def _visibility(self, value) -> bool | None:
        text = self.normalizer.text(value)
        return self.visibility_aliases.get(text.lower()) if text else True

    def _image_urls(self, primary, gallery) -> list[str]:
        values = []
        for item in (primary, gallery):
            text = self.normalizer.text(item)
            if text:
                values.extend([part.strip() for part in split(r"[;,\n\r\t ]+", text) if part.strip()])
        return list(dict.fromkeys(values))

    def _valid_url(self, url: str) -> bool:
        return url.startswith("https://") or url.startswith("http://")



class HistoricalImportService:
    order_status_aliases = {
        "new": OrderStatus.NEW, "нове": OrderStatus.NEW, "новий": OrderStatus.NEW, "нове замовлення": OrderStatus.NEW, "новий заказ": OrderStatus.NEW,
        "confirmed": OrderStatus.CONFIRMED, "підтверджено": OrderStatus.CONFIRMED, "подтверждено": OrderStatus.CONFIRMED, "підтверджений": OrderStatus.CONFIRMED,
        "shipped": OrderStatus.SHIPPED, "відправлено": OrderStatus.SHIPPED, "отправлено": OrderStatus.SHIPPED,
        "delivered": OrderStatus.DELIVERED, "доставлено": OrderStatus.DELIVERED, "отримано": OrderStatus.DELIVERED, "получено": OrderStatus.DELIVERED,
        "completed": OrderStatus.COMPLETED, "завершено": OrderStatus.COMPLETED, "виконано": OrderStatus.COMPLETED,
        "returned": OrderStatus.RETURNED, "повернення": OrderStatus.RETURNED, "повернено": OrderStatus.RETURNED, "возврат": OrderStatus.RETURNED,
        "cancelled": OrderStatus.CANCELLED, "скасовано": OrderStatus.CANCELLED, "отменено": OrderStatus.CANCELLED, "відміна": OrderStatus.CANCELLED,
    }
    payment_status_aliases = {
        "pending": PaymentStatus.PENDING, "очікує": PaymentStatus.PENDING, "не оплачено": PaymentStatus.PENDING,
        "paid": PaymentStatus.PAID, "оплачено": PaymentStatus.PAID,
        "cod": PaymentStatus.COD, "накладений платіж": PaymentStatus.COD, "післяплата": PaymentStatus.COD, "наложка": PaymentStatus.COD,
        "refunded": PaymentStatus.REFUNDED, "повернено оплату": PaymentStatus.REFUNDED, "refund": PaymentStatus.REFUNDED,
    }
    platform_aliases = {item.value.lower(): item for item in AdCampaignPlatform}

    def __init__(self, db: Session, lookup: ImportEntityLookupRepository) -> None:
        self.db = db
        self.lookup = lookup
        self.normalizer = ExcelValueNormalizer()

    def dry_run(self, workspace_id: UUID, job_id: UUID, entity_type: str, sheet_name: str, rows: list[dict], mapping: dict[str, str]) -> ImportReportResponse:
        if entity_type == "orders_history":
            analysis = self._analyze_orders(workspace_id, rows, mapping)
            return self._report(job_id, entity_type, sheet_name, rows, analysis["issues"], analysis["metrics"])
        analysis = self._analyze_advertising(workspace_id, rows, mapping)
        return self._report(job_id, entity_type, sheet_name, rows, analysis["issues"], analysis["metrics"])

    def import_rows(self, workspace_id: UUID, entity_type: str, rows: list[dict], mapping: dict[str, str], options: dict | None, actor_user_id: UUID | None = None) -> list[tuple[int | None, ImportJobLogStatus, str]]:
        if entity_type == "orders_history":
            return self._import_orders(workspace_id, rows, mapping, bool((options or {}).get("affect_inventory", False)), actor_user_id)
        return self._import_advertising(workspace_id, rows, mapping, actor_user_id)

    def _mapped(self, row: dict, mapping: dict[str, str]) -> dict:
        return MappingValidationService().normalized_row(map_row(row, mapping))

    def _order_key(self, data: dict, row_number: int) -> str:
        order_number = self.normalizer.text(data.get("order_number"))
        if order_number:
            return order_number
        order_date = data.get("order_date")
        return f"IMPORT-{order_date.date().isoformat() if order_date else 'NO-DATE'}-ROW-{row_number}"

    def _group_order_rows(self, rows: list[dict], mapping: dict[str, str]) -> dict[str, list[tuple[int, dict]]]:
        grouped: dict[str, list[tuple[int, dict]]] = defaultdict(list)
        for row_number, row in enumerate(rows, start=2):
            data = self._mapped(row, mapping)
            grouped[self._order_key(data, row_number)].append((row_number, data))
        return grouped

    def _analyze_orders(self, workspace_id: UUID, rows: list[dict], mapping: dict[str, str]) -> dict:
        issues: list[ImportValidationIssue] = []
        groups = self._group_order_rows(rows, mapping)
        metrics = defaultdict(int)
        estimated_revenue = Decimal("0")
        estimated_ad_cost = Decimal("0")
        estimated_profit = Decimal("0")
        customer_keys: set[str] = set()
        matched_customers = set()
        variants_missing = set()
        for order_key, items in groups.items():
            if self.lookup.find_order_by_number(workspace_id, order_key):
                for row_number, _data in items:
                    issues.append(issue(row_number, "WARNING", "order_number", "Duplicate order skipped"))
                metrics["duplicate_orders"] += 1
                continue
            order_revenue = Decimal("0")
            order_cost = Decimal("0")
            first = items[0][1]
            if self._find_customer(workspace_id, first):
                matched_customers.add(order_key)
            elif first.get("customer_phone") or first.get("instagram_username") or first.get("customer_name"):
                customer_keys.add(str(first.get("customer_phone") or first.get("instagram_username") or first.get("customer_name")).lower())
            for row_number, data in items:
                sku = data.get("variant_sku")
                variant = self.lookup.find_variant(workspace_id, sku=sku)
                if not sku:
                    issues.append(issue(row_number, "ERROR", "variant_sku", "Product variant SKU is required."))
                    variants_missing.add(row_number)
                    continue
                if variant is None:
                    issues.append(issue(row_number, "ERROR", "variant_sku", "Product variant not found. Import product catalog first."))
                    variants_missing.add(row_number)
                    continue
                qty = int(data.get("quantity") or 0)
                if qty <= 0:
                    issues.append(issue(row_number, "ERROR", "quantity", "Quantity must be greater than 0."))
                    continue
                unit_price = data.get("unit_price")
                if unit_price is None and data.get("order_total") is not None:
                    unit_price = Decimal(data["order_total"]) / Decimal(qty)
                if unit_price is None or unit_price < 0:
                    issues.append(issue(row_number, "ERROR", "unit_price", "Unit price is required."))
                    continue
                unit_cost = data.get("unit_cost")
                if unit_cost is None:
                    unit_cost = Decimal(str(getattr(variant, "purchase_price", 0) or 0))
                    issues.append(issue(row_number, "WARNING", "unit_cost", "Unit cost missing; using catalog cost or zero."))
                line_total = Decimal(unit_price) * qty
                line_cost = Decimal(unit_cost) * qty
                order_revenue += line_total
                order_cost += line_cost
                metrics["order_items_detected"] += 1
                metrics["variants_matched"] += 1
            if first.get("tracking_number"):
                metrics["shipments_detected"] += 1
            if first.get("order_total") is not None and abs(Decimal(first["order_total"]) - order_revenue) > Decimal("0.01"):
                issues.append(issue(items[0][0], "WARNING", "order_total", "Order total differs from item totals."))
            order_costs = sum((Decimal(first.get(field) or 0) for field in ("ad_cost", "shipping_cost", "cod_fee", "other_cost")), Decimal("0"))
            estimated_revenue += order_revenue
            estimated_ad_cost += Decimal(first.get("ad_cost") or 0)
            estimated_profit += order_revenue - order_cost - order_costs
        metrics.update({
            "orders_detected": len(groups),
            "customers_matched": len(matched_customers),
            "customers_to_create": len(customer_keys),
            "variants_missing": len(variants_missing),
            "ready_orders": max(len(groups) - metrics["duplicate_orders"], 0),
            "ready_items": metrics["order_items_detected"],
            "estimated_revenue": estimated_revenue,
            "estimated_ad_cost": estimated_ad_cost,
            "estimated_profit": estimated_profit,
        })
        return {"issues": issues, "metrics": dict(metrics)}

    def _find_customer(self, workspace_id: UUID, data: dict) -> Customer | None:
        customer = self.lookup.find_customer(workspace_id, data.get("customer_phone"), data.get("instagram_username"))
        if customer or not data.get("customer_name"):
            return customer
        return self.db.execute(select(Customer).where(Customer.workspace_id == workspace_id, Customer.deleted_at.is_(None), Customer.name == data.get("customer_name"))).scalar_one_or_none()

    def _get_or_create_customer(self, workspace_id: UUID, data: dict) -> Customer | None:
        customer = self._find_customer(workspace_id, data)
        if customer:
            return customer
        if not (data.get("customer_name") or data.get("customer_phone") or data.get("instagram_username")):
            return None
        customer = Customer(workspace_id=workspace_id, name=data.get("customer_name") or data.get("customer_phone") or "Historical customer", phone=data.get("customer_phone"), instagram_username=data.get("instagram_username"))
        self.db.add(customer); self.db.flush()
        return customer

    def _normalize_order_status(self, value) -> tuple[OrderStatus, bool]:
        text = (self.normalizer.text(value) or "").lower()
        return self.order_status_aliases.get(text, OrderStatus.NEW), bool(text and text not in self.order_status_aliases)

    def _normalize_payment_status(self, value) -> tuple[PaymentStatus, bool]:
        text = (self.normalizer.text(value) or "").lower()
        return self.payment_status_aliases.get(text, PaymentStatus.PENDING), bool(text and text not in self.payment_status_aliases)

    def _shipment_status_for_order(self, status: OrderStatus) -> str:
        if status == OrderStatus.SHIPPED:
            return ShipmentStatus.IN_TRANSIT.value
        if status in {OrderStatus.DELIVERED, OrderStatus.COMPLETED}:
            return ShipmentStatus.DELIVERED.value
        if status == OrderStatus.RETURNED:
            return ShipmentStatus.RETURNED.value
        return ShipmentStatus.DRAFT.value

    def _import_orders(self, workspace_id: UUID, rows: list[dict], mapping: dict[str, str], affect_inventory: bool, actor_user_id: UUID | None) -> list[tuple[int | None, ImportJobLogStatus, str]]:
        analysis = self._analyze_orders(workspace_id, rows, mapping)
        if any(item.severity == "ERROR" for item in analysis["issues"]):
            return [(item.row_number, ImportJobLogStatus.FAILED, item.message) for item in analysis["issues"] if item.severity == "ERROR"]
        results: list[tuple[int | None, ImportJobLogStatus, str]] = []
        order_service = OrderService(self.db)
        for order_key, grouped_rows in self._group_order_rows(rows, mapping).items():
            if self.lookup.find_order_by_number(workspace_id, order_key):
                for row_number, _data in grouped_rows:
                    results.append((row_number, ImportJobLogStatus.SKIPPED, "Duplicate order skipped"))
                continue
            first = grouped_rows[0][1]
            customer = self._get_or_create_customer(workspace_id, first)
            order_status, status_warn = self._normalize_order_status(first.get("order_status"))
            payment_status, payment_warn = self._normalize_payment_status(first.get("payment_status"))
            item_payloads = []
            for row_number, data in grouped_rows:
                variant = self.lookup.find_variant(workspace_id, sku=data.get("variant_sku"))
                qty = int(data.get("quantity") or 0)
                unit_price = data.get("unit_price") or ((Decimal(data["order_total"]) / Decimal(qty)) if data.get("order_total") is not None and qty > 0 else Decimal("0"))
                unit_cost = data.get("unit_cost") if data.get("unit_cost") is not None else Decimal(str(getattr(variant, "purchase_price", 0) or 0))
                item_payloads.append(OrderItemCreate(product_variant_id=variant.id, quantity=qty, unit_price=unit_price, unit_cost=unit_cost))
            created_at = first.get("order_date")
            payload = OrderCreate(customer_id=customer.id if customer else None, status=order_status, payment_status=payment_status, is_historical=True, items=item_payloads, ad_cost=first.get("ad_cost") or Decimal("0"), shipping_cost=first.get("shipping_cost") or Decimal("0"), cod_fee=first.get("cod_fee") or Decimal("0"), other_cost=first.get("other_cost") or Decimal("0"), notes=first.get("notes"))
            try:
                order = order_service.create(workspace_id, payload, actor_user_id, affect_inventory=affect_inventory, order_number=order_key, created_at=created_at, completed_at=created_at if order_status in {OrderStatus.COMPLETED, OrderStatus.DELIVERED} else None)
            except OrderServiceError as exc:
                results.append((grouped_rows[0][0], ImportJobLogStatus.FAILED, str(exc))); continue
            if first.get("tracking_number"):
                carrier = first.get("carrier") if first.get("carrier") in {item.value for item in ShipmentCarrier} else ShipmentCarrier.NOVA_POSHTA.value
                shipment = Shipment(workspace_id=workspace_id, order_id=order.id, customer_id=order.customer_id, tracking_number=first.get("tracking_number"), carrier=carrier, status=self._shipment_status_for_order(order_status), city=first.get("city"), warehouse=first.get("warehouse"), shipping_cost=first.get("shipping_cost"), notes="Historical import")
                self.db.add(shipment); self.db.flush()
            status = ImportJobLogStatus.WARNING if status_warn or payment_warn or customer is None else ImportJobLogStatus.SUCCESS
            results.append((grouped_rows[0][0], status, "Historical order imported" if status == ImportJobLogStatus.SUCCESS else "Historical order imported with warnings"))
            for row_number, _data in grouped_rows[1:]:
                results.append((row_number, ImportJobLogStatus.SUCCESS, "Historical order item imported"))
        return results

    def _normalize_platform(self, value) -> tuple[AdCampaignPlatform, bool]:
        text = (self.normalizer.text(value) or "instagram").lower()
        return self.platform_aliases.get(text, AdCampaignPlatform.OTHER), text not in self.platform_aliases

    def _find_campaign(self, workspace_id: UUID, name: str | None, platform: AdCampaignPlatform) -> AdCampaign | None:
        if not name:
            return None
        return self.db.execute(select(AdCampaign).where(AdCampaign.workspace_id == workspace_id, AdCampaign.deleted_at.is_(None), AdCampaign.name == name, AdCampaign.platform == platform.value)).scalar_one_or_none()

    def _analyze_advertising(self, workspace_id: UUID, rows: list[dict], mapping: dict[str, str]) -> dict:
        issues: list[ImportValidationIssue] = []
        metrics = defaultdict(int)
        campaigns = set()
        spend = Decimal("0"); revenue = Decimal("0"); profit = Decimal("0")
        for row_number, row in enumerate(rows, start=2):
            data = self._mapped(row, mapping)
            if not data.get("campaign_name"):
                issues.append(issue(row_number, "ERROR", "campaign_name", "Campaign name is required.")); continue
            if not data.get("metric_date"):
                issues.append(issue(row_number, "ERROR", "metric_date", "Metric date is required.")); continue
            if data.get("spend") is not None and data["spend"] < 0:
                issues.append(issue(row_number, "ERROR", "spend", "Spend cannot be negative.")); continue
            platform, platform_warn = self._normalize_platform(data.get("platform"))
            if platform_warn:
                issues.append(issue(row_number, "WARNING", "platform", "Unknown platform mapped to OTHER."))
            campaign_key = (data["campaign_name"], platform.value)
            campaigns.add(campaign_key)
            campaign = self._find_campaign(workspace_id, data["campaign_name"], platform)
            if campaign:
                metrics["campaigns_reused"] += 1
                if self.lookup.find_ad_metric_by_campaign_date(workspace_id, campaign.id, data["metric_date"].date()):
                    issues.append(issue(row_number, "WARNING", "metric_date", "Duplicate ad metric skipped")); metrics["duplicate_metrics"] += 1
            metrics["metrics_detected"] += 1
            spend += data.get("spend") or Decimal("0"); revenue += data.get("revenue") or Decimal("0"); profit += data.get("net_profit") or Decimal("0")
        metrics.update({"campaigns_detected": len(campaigns), "campaigns_to_create": len([c for c in campaigns if not self._find_campaign(workspace_id, c[0], AdCampaignPlatform(c[1]))]), "estimated_spend": spend, "estimated_revenue": revenue, "estimated_net_profit": profit, "estimated_roas": (revenue / spend if spend else Decimal("0"))})
        return {"issues": issues, "metrics": dict(metrics)}

    def _import_advertising(self, workspace_id: UUID, rows: list[dict], mapping: dict[str, str], actor_user_id: UUID | None) -> list[tuple[int | None, ImportJobLogStatus, str]]:
        analysis = self._analyze_advertising(workspace_id, rows, mapping)
        if any(item.severity == "ERROR" for item in analysis["issues"]):
            return [(item.row_number, ImportJobLogStatus.FAILED, item.message) for item in analysis["issues"] if item.severity == "ERROR"]
        campaign_service = AdCampaignService(self.db); metric_service = AdMetricService(self.db)
        results = []
        for row_number, row in enumerate(rows, start=2):
            data = self._mapped(row, mapping)
            platform, platform_warn = self._normalize_platform(data.get("platform"))
            campaign = self._find_campaign(workspace_id, data.get("campaign_name"), platform)
            if campaign is None:
                campaign = campaign_service.create(workspace_id, AdCampaignCreate(name=data["campaign_name"], platform=platform, status=AdCampaignStatus.ACTIVE, objective=AdCampaignObjective.SALES, budget_type=AdCampaignBudgetType.DAILY, start_date=data["metric_date"].date() if data.get("metric_date") else None), actor_user_id)
            if self.lookup.find_ad_metric_by_campaign_date(workspace_id, campaign.id, data["metric_date"].date()):
                results.append((row_number, ImportJobLogStatus.SKIPPED, "Duplicate ad metric skipped")); continue
            try:
                metric_service.create(workspace_id, AdMetricCreate(campaign_id=campaign.id, metric_date=data["metric_date"].date(), spend=data.get("spend") or Decimal("0"), impressions=int(data.get("impressions") or 0), reach=int(data.get("reach") or 0), clicks=int(data.get("clicks") or 0), messages=int(data.get("messages") or 0), leads=int(data.get("leads") or 0), orders=int(data.get("orders") or 0), revenue=data.get("revenue") or Decimal("0"), net_profit=data.get("net_profit") or Decimal("0")), actor_user_id)
            except AdvertisingServiceError as exc:
                results.append((row_number, ImportJobLogStatus.FAILED, str(exc))); continue
            results.append((row_number, ImportJobLogStatus.WARNING if platform_warn else ImportJobLogStatus.SUCCESS, "Historical ad metric imported" if not platform_warn else "Historical ad metric imported with warnings"))
        return results

    def _report(self, job_id: UUID, entity_type: str, sheet_name: str, rows: list[dict], issues: list[ImportValidationIssue], metrics: dict) -> ImportReportResponse:
        report = import_report(job_id, entity_type, sheet_name, rows, report_from_issues(rows, issues), {key: value for key, value in metrics.items() if isinstance(value, int)})
        for field in ("orders_detected", "order_items_detected", "customers_matched", "customers_to_create", "variants_matched", "variants_missing", "shipments_detected", "duplicate_orders", "ready_orders", "ready_items", "campaigns_detected", "campaigns_to_create", "campaigns_reused", "metrics_detected", "duplicate_metrics"):
            setattr(report, field, int(metrics.get(field, 0)))
        for field in ("estimated_revenue", "estimated_ad_cost", "estimated_profit", "estimated_spend", "estimated_net_profit", "estimated_roas"):
            setattr(report, field, metrics.get(field, Decimal("0")))
        if entity_type == "orders_history":
            report.estimated_entities_to_create = int(metrics.get("ready_orders", 0)) + int(metrics.get("ready_items", 0)) + int(metrics.get("customers_to_create", 0)) + int(metrics.get("shipments_detected", 0))
            report.ready_to_import_rows = int(metrics.get("ready_items", 0))
            report.duplicate_rows = int(metrics.get("duplicate_orders", 0))
        else:
            report.estimated_entities_to_create = int(metrics.get("campaigns_to_create", 0)) + int(metrics.get("metrics_detected", 0))
            report.duplicate_rows = int(metrics.get("duplicate_metrics", 0))
        return report


class EntityImportService:
    def __init__(self, db: Session) -> None:
        self.db = db
        self.lookup = ImportEntityLookupRepository(db)
        self.validator = MappingValidationService()

    def import_row(self, workspace_id: UUID, entity_type: str, raw_row: dict, column_mapping: dict[str, str]) -> tuple[ImportJobLogStatus, str]:
        data = self.validator.normalized_row(map_row(raw_row, column_mapping))
        if entity_type == "customers":
            return self._customer(workspace_id, data)
        if entity_type == "products":
            return self._product(workspace_id, data)
        if entity_type == "product_variants":
            return self._variant(workspace_id, data)
        if entity_type == "inventory":
            return self._inventory(workspace_id, data)
        if entity_type == "orders":
            return self._order(workspace_id, data)
        if entity_type == "ad_campaigns":
            return self._ad_campaign(workspace_id, data)
        if entity_type == "ad_metrics":
            return self._ad_metric(workspace_id, data)
        if entity_type == "shipments":
            return self._shipment(workspace_id, data)
        return ImportJobLogStatus.FAILED, "Unsupported entity type"

    def _customer(self, workspace_id: UUID, data: dict) -> tuple[ImportJobLogStatus, str]:
        if not any(data.get(field) for field in ("name", "phone", "instagram_username")):
            return ImportJobLogStatus.FAILED, "Customer requires name, phone, or instagram_username"
        if self.lookup.find_customer(workspace_id, data.get("phone"), data.get("instagram_username")):
            return ImportJobLogStatus.SKIPPED, "Duplicate customer skipped"
        self.db.add(Customer(workspace_id=workspace_id, name=data.get("name") or data.get("phone") or data.get("instagram_username"), phone=data.get("phone"), instagram_username=data.get("instagram_username"), city=data.get("city"), region=data.get("region")))
        self.db.flush()
        return ImportJobLogStatus.SUCCESS, "Customer created"

    def _product(self, workspace_id: UUID, data: dict) -> tuple[ImportJobLogStatus, str]:
        if not data.get("name"):
            return ImportJobLogStatus.FAILED, "Product requires name"
        if self.lookup.find_product_by_sku(workspace_id, data.get("sku")):
            return ImportJobLogStatus.SKIPPED, "Duplicate product skipped"
        self.db.add(Product(workspace_id=workspace_id, name=data["name"], sku=data.get("sku"), description=data.get("description")))
        self.db.flush()
        return ImportJobLogStatus.SUCCESS, "Product created"

    def _variant(self, workspace_id: UUID, data: dict) -> tuple[ImportJobLogStatus, str]:
        product = self.lookup.find_product_by_name_or_sku(workspace_id, data.get("product_name"), data.get("product_sku"))
        if product is None:
            return ImportJobLogStatus.FAILED, "Product not found for variant"
        sku = data.get("variant_sku")
        if self.lookup.find_variant(workspace_id, sku=sku, product_id=product.id, color=data.get("color"), size=data.get("size")):
            return ImportJobLogStatus.SKIPPED, "Duplicate variant skipped"
        variant = ProductVariant(workspace_id=workspace_id, product_id=product.id, sku=sku or f"{product.sku or product.name}-{data.get('color') or ''}-{data.get('size') or ''}", color=data.get("color"), size=data.get("size"), price=data.get("selling_price"))
        self.db.add(variant); self.db.flush()
        self.db.add(Inventory(workspace_id=workspace_id, product_variant_id=variant.id, stock_quantity=0, reserved_quantity=0, minimum_quantity=0)); self.db.flush()
        return ImportJobLogStatus.SUCCESS, "Variant created"

    def _inventory(self, workspace_id: UUID, data: dict) -> tuple[ImportJobLogStatus, str]:
        variant = self.lookup.find_variant(workspace_id, sku=data.get("variant_sku"))
        if variant is None:
            return ImportJobLogStatus.FAILED, "Variant not found for inventory"
        inventory = self.lookup.inventory_by_variant(workspace_id, variant.id)
        if inventory is None:
            inventory = Inventory(workspace_id=workspace_id, product_variant_id=variant.id)
            self.db.add(inventory)
        inventory.stock_quantity = int(data.get("stock_quantity") or 0)
        inventory.reserved_quantity = int(data.get("reserved_quantity") or 0)
        inventory.minimum_quantity = int(data.get("minimum_quantity") or 0)
        self.db.flush()
        return ImportJobLogStatus.SUCCESS, "Inventory updated"

    def _order(self, workspace_id: UUID, data: dict) -> tuple[ImportJobLogStatus, str]:
        customer = self.lookup.find_customer(workspace_id, data.get("customer_phone"), data.get("instagram_username"))
        revenue = data.get("revenue") or data.get("order_total")
        created_at = data.get("created_at") or data.get("order_date")
        if revenue is None or created_at is None:
            return ImportJobLogStatus.FAILED, "Order requires revenue/order_total and created_at/order_date"
        order = Order(workspace_id=workspace_id, order_number=f"IMP-{datetime.now(UTC).strftime('%Y%m%d%H%M%S%f')}", customer_id=customer.id if customer else None, status=OrderStatus.COMPLETED.value, payment_status=PaymentStatus.PAID.value, revenue=revenue, product_cost=Decimal("0"), ad_cost=data.get("ad_cost") or Decimal("0"), shipping_cost=data.get("shipping_cost") or Decimal("0"), cod_fee=data.get("cod_fee") or Decimal("0"), other_cost=data.get("other_cost") or Decimal("0"), net_profit=data.get("net_profit") or revenue, created_at=created_at, completed_at=created_at)
        self.db.add(order); self.db.flush()
        return ImportJobLogStatus.WARNING if customer is None else ImportJobLogStatus.SUCCESS, "Order created" if customer else "Order created without matched customer"

    def _ad_campaign(self, workspace_id: UUID, data: dict) -> tuple[ImportJobLogStatus, str]:
        if not data.get("name"):
            return ImportJobLogStatus.FAILED, "Ad campaign requires name"
        if self.lookup.find_ad_campaign_by_name(workspace_id, data.get("name")):
            return ImportJobLogStatus.SKIPPED, "Duplicate ad campaign skipped"
        campaign = AdCampaign(workspace_id=workspace_id, name=data["name"], platform=(data.get("platform") or "INSTAGRAM"), objective=(data.get("objective") or "MESSAGES"), budget_type=(data.get("budget_type") or "MANUAL"), daily_budget=data.get("daily_budget"), total_budget=data.get("total_budget"), start_date=data.get("start_date").date() if data.get("start_date") else None, end_date=data.get("end_date").date() if data.get("end_date") else None, notes=data.get("notes"))
        self.db.add(campaign); self.db.flush()
        return ImportJobLogStatus.SUCCESS, "Ad campaign created"

    def _ad_metric(self, workspace_id: UUID, data: dict) -> tuple[ImportJobLogStatus, str]:
        campaign = self.lookup.find_ad_campaign_by_id(workspace_id, data.get("campaign_id")) or self.lookup.find_ad_campaign_by_name(workspace_id, data.get("campaign_name"))
        if campaign is None:
            return ImportJobLogStatus.FAILED, "Ad campaign not found for metric"
        metric_date = data.get("metric_date")
        if metric_date is None:
            return ImportJobLogStatus.FAILED, "Ad metric requires metric_date"
        metric_date_value = metric_date.date()
        if self.lookup.find_ad_metric_by_campaign_date(workspace_id, campaign.id, metric_date_value):
            return ImportJobLogStatus.SKIPPED, "Duplicate ad metric skipped"
        metric = AdMetric(workspace_id=workspace_id, campaign_id=campaign.id, metric_date=metric_date_value, spend=data.get("spend") or Decimal("0"), impressions=int(data.get("impressions") or 0), reach=int(data.get("reach") or 0), clicks=int(data.get("clicks") or 0), messages=int(data.get("messages") or 0), leads=int(data.get("leads") or 0), orders=int(data.get("orders") or 0), revenue=data.get("revenue") or Decimal("0"), net_profit=data.get("net_profit") or Decimal("0"))
        self.db.add(metric); self.db.flush()
        return ImportJobLogStatus.SUCCESS, "Ad metric created"

    def _shipment(self, workspace_id: UUID, data: dict) -> tuple[ImportJobLogStatus, str]:
        order = self.lookup.find_order_by_number(workspace_id, data.get("order_number"))
        if order is None:
            return ImportJobLogStatus.FAILED, "Order not found for shipment"
        if self.lookup.find_shipment_by_order(workspace_id, order.id):
            return ImportJobLogStatus.SKIPPED, "Active shipment already exists for order"
        if self.lookup.find_shipment_by_tracking(workspace_id, data.get("tracking_number")):
            return ImportJobLogStatus.SKIPPED, "Duplicate shipment tracking number skipped"
        status_value = data.get("status") or ShipmentStatus.DRAFT.value
        status_value = status_value if status_value in {item.value for item in ShipmentStatus} else ShipmentStatus.DRAFT.value
        if status_value != ShipmentStatus.DRAFT.value and not data.get("tracking_number"):
            return ImportJobLogStatus.FAILED, "tracking_number is required for non-draft shipment"
        carrier_value = data.get("carrier") or ShipmentCarrier.NOVA_POSHTA.value
        carrier_value = carrier_value if carrier_value in {item.value for item in ShipmentCarrier} else ShipmentCarrier.NOVA_POSHTA.value
        shipment = Shipment(workspace_id=workspace_id, order_id=order.id, customer_id=order.customer_id, tracking_number=data.get("tracking_number"), carrier=carrier_value, status=status_value, recipient_name=data.get("recipient_name"), recipient_phone=data.get("recipient_phone"), city=data.get("city"), warehouse=data.get("warehouse"), shipping_cost=data.get("shipping_cost"), cod_amount=data.get("cod_amount"), declared_value=data.get("declared_value"), notes=data.get("notes"))
        self.db.add(shipment); self.db.flush()
        return ImportJobLogStatus.SUCCESS, "Shipment created"


class ImportService:
    def __init__(self, db: Session) -> None:
        self.db = db
        self.jobs = ImportJobRepository(db)
        self.logs = ImportJobLogRepository(db)
        self.parser = ExcelParserService()
        self.validator = MappingValidationService()
        self.suggestions = MappingSuggestionService()
        self.entity_importer = EntityImportService(db)
        self.product_catalog_importer = ProductCatalogImportService(db)
        self.lookup = ImportEntityLookupRepository(db)
        self.historical_importer = HistoricalImportService(db, self.lookup)
        self.audit_logs = AuditLogRepository(db)

    async def upload(self, workspace_id: UUID, file: UploadFile, actor_user_id: UUID | None) -> ImportJob:
        safe_name = self._safe_filename(file.filename or "import.xlsx")
        if not safe_name.lower().endswith(".xlsx"):
            raise ImportServiceError("Only .xlsx files are supported")
        content = await file.read()
        if len(content) > get_settings().import_max_file_size_mb * 1024 * 1024:
            raise ImportServiceError("Import file exceeds size limit")
        job = self.jobs.create(ImportJob(workspace_id=workspace_id, file_name=safe_name, file_type="xlsx", file_path="pending", status=ImportJobStatus.UPLOADED.value, created_by=actor_user_id))
        directory = Path(get_settings().import_storage_path) / str(workspace_id) / str(job.id)
        directory.mkdir(parents=True, exist_ok=True)
        path = directory / safe_name
        path.write_bytes(content)
        job.file_path = str(path)
        self.audit_logs.create(workspace_id=workspace_id, user_id=actor_user_id, entity_type="ImportJob", entity_id=job.id, action="IMPORT_UPLOAD", new_value=safe_job_snapshot(job))
        self.db.commit(); self.db.refresh(job); return job

    def list_sheets(self, workspace_id: UUID, job_id: UUID) -> list[str]:
        return self.parser.list_sheets(self._job(workspace_id, job_id).file_path)

    def preview(self, workspace_id: UUID, job_id: UUID, sheet_name: str, limit: int) -> tuple[list[str], list[dict]]:
        job = self._job(workspace_id, job_id)
        columns, rows = self.parser.preview(job.file_path, sheet_name, limit)
        job.status = ImportJobStatus.PREVIEWED.value
        self.db.commit()
        return columns, rows

    def suggest_mapping(self, workspace_id: UUID, job_id: UUID, sheet_name: str, entity_type: str) -> SuggestMappingResponse:
        job = self._job(workspace_id, job_id)
        columns, _rows = self.parser.preview(job.file_path, sheet_name, 1)
        if entity_type == "product_catalog":
            mapping = {field: column for field, column in PRODUCT_CATALOG_MAPPING.items() if column in columns}
            required = [field for field in ("product_sku", "product_sku_fallback", "product_name", "product_name_fallback", "variant_sku", "selling_price") if field not in mapping]
            return SuggestMappingResponse(suggested_mapping=mapping, confidence={field: 1.0 for field in mapping}, unmapped_columns=[column for column in columns if column not in mapping.values()], required_fields_missing=required)
        if entity_type in {"orders_history", "advertising_history"}:
            preset = ORDERS_HISTORY_MAPPING if entity_type == "orders_history" else ADVERTISING_HISTORY_MAPPING
            mapping = {field: column for field, aliases in FIELD_ALIASES[entity_type].items() for column in columns if normalize_name(column) in {normalize_name(alias) for alias in aliases} or column == preset.get(field)}
            return SuggestMappingResponse(suggested_mapping=mapping, confidence={field: 1.0 for field in mapping}, unmapped_columns=[column for column in columns if column not in mapping.values()], required_fields_missing=[])
        return self.suggestions.suggest(columns, entity_type)

    def dry_run(self, workspace_id: UUID, job_id: UUID, entity_type: str, sheet_name: str, column_mapping: dict[str, str], actor_user_id: UUID | None = None, options: dict | None = None) -> ImportReportResponse:
        job = self._job(workspace_id, job_id)
        _columns, rows = self.parser.read_rows(job.file_path, sheet_name)
        if entity_type in {"orders_history", "advertising_history"}:
            response = self.historical_importer.dry_run(workspace_id, job.id, entity_type, sheet_name, rows, column_mapping)
        else:
            report = self.product_catalog_importer.validate(rows, column_mapping, workspace_id) if entity_type == "product_catalog" else self.validator.validate(entity_type, column_mapping, rows, workspace_id, self.lookup)
            response = import_report(job.id, entity_type, sheet_name, rows, report, self.product_catalog_importer.metrics(rows, column_mapping, report) if entity_type == "product_catalog" else None)
        self.audit_logs.create(workspace_id=workspace_id, user_id=actor_user_id, entity_type="ImportJob", entity_id=job.id, action="IMPORT_DRY_RUN", new_value=response.model_dump(mode="json", exclude={"sample_errors", "sample_warnings"}))
        self.db.commit()
        return response

    def validate(self, workspace_id: UUID, job_id: UUID, entity_type: str, sheet_name: str, column_mapping: dict[str, str], actor_user_id: UUID | None, options: dict | None = None) -> ImportValidationReport:
        job = self._job(workspace_id, job_id)
        _columns, rows = self.parser.read_rows(job.file_path, sheet_name)
        report = self.product_catalog_importer.validate(rows, column_mapping, workspace_id) if entity_type == "product_catalog" else self.validator.validate(entity_type, column_mapping, rows, workspace_id, self.lookup)
        job.total_rows = report.total_rows
        if report.is_valid:
            job.status = ImportJobStatus.VALIDATED.value
        self.audit_logs.create(workspace_id=workspace_id, user_id=actor_user_id, entity_type="ImportJob", entity_id=job.id, action="IMPORT_VALIDATE", new_value={"is_valid": report.is_valid, "total_rows": report.total_rows, "errors": len(report.errors), "warnings": len(report.warnings)})
        self.db.commit()
        return report

    def execute(self, workspace_id: UUID, job_id: UUID, entity_type: str, sheet_name: str, column_mapping: dict[str, str], mode: str, actor_user_id: UUID | None, dry_run: bool = False, options: dict | None = None):
        if dry_run:
            return self.dry_run(workspace_id, job_id, entity_type, sheet_name, column_mapping, actor_user_id, options)
        if mode != "create_only":
            raise ImportServiceError("Only create_only import mode is supported")
        job = self._job(workspace_id, job_id)
        _columns, rows = self.parser.read_rows(job.file_path, sheet_name)
        if entity_type in {"orders_history", "advertising_history"}:
            dry_report = self.historical_importer.dry_run(workspace_id, job.id, entity_type, sheet_name, rows, column_mapping)
            report = ImportValidationReport(is_valid=dry_report.error_rows == 0, total_rows=len(rows), errors=[item.message for item in dry_report.sample_errors], warnings=[item.message for item in dry_report.sample_warnings], issues=[*dry_report.sample_errors, *dry_report.sample_warnings])
        else:
            report = self.product_catalog_importer.validate(rows, column_mapping, workspace_id) if entity_type == "product_catalog" else self.validator.validate(entity_type, column_mapping, rows, workspace_id, self.lookup)
        if any(item.severity == "ERROR" for item in report.issues):
            job.status = ImportJobStatus.FAILED.value
            job.failed_rows = len(rows)
            self.audit_logs.create(workspace_id=workspace_id, user_id=actor_user_id, entity_type="ImportJob", entity_id=job.id, action="IMPORT_FAILED", new_value={"total_rows": len(rows), "errors": len(report.errors)})
            self.db.commit()
            raise ImportServiceError("Import mapping is invalid")
        job.status = ImportJobStatus.IMPORTING.value
        job.total_rows = len(rows)
        job.processed_rows = job.processed_rows or 0
        job.success_rows = job.success_rows or 0
        job.failed_rows = job.failed_rows or 0
        self.audit_logs.create(workspace_id=workspace_id, user_id=actor_user_id, entity_type="ImportJob", entity_id=job.id, action="IMPORT_EXECUTE", new_value={"entity_type": entity_type, "rows": len(rows)})
        import_results = self.product_catalog_importer.import_rows(workspace_id, rows, column_mapping) if entity_type == "product_catalog" else (self.historical_importer.import_rows(workspace_id, entity_type, rows, column_mapping, options, actor_user_id) if entity_type in {"orders_history", "advertising_history"} else None)
        for index, row in enumerate(rows, start=2):
            if import_results is not None:
                if index - 2 >= len(import_results):
                    continue
                row_number, status, message = import_results[index - 2]
                raw_data = None
                log_row_number = row_number or index
            else:
                status, message = self.entity_importer.import_row(workspace_id, entity_type, row, column_mapping)
                raw_data = row
                log_row_number = index
            self.logs.create(ImportJobLog(workspace_id=workspace_id, import_job_id=job.id, row_number=log_row_number, entity_type=entity_type, status=status.value, message=message, raw_data=raw_data))
            job.processed_rows += 1
            if status == ImportJobLogStatus.SUCCESS:
                job.success_rows += 1
            elif status == ImportJobLogStatus.FAILED:
                job.failed_rows += 1
        job.completed_at = datetime.now(UTC)
        job.status = ImportJobStatus.COMPLETED.value if job.failed_rows == 0 else (ImportJobStatus.FAILED.value if job.success_rows == 0 else ImportJobStatus.PARTIALLY_COMPLETED.value)
        self.audit_logs.create(workspace_id=workspace_id, user_id=actor_user_id, entity_type="ImportJob", entity_id=job.id, action="IMPORT_COMPLETED" if job.success_rows else "IMPORT_FAILED", new_value=safe_job_snapshot(job))
        self.db.commit(); self.db.refresh(job); return job

    def list_logs(self, workspace_id: UUID, job_id: UUID, status: str | None = None) -> list[ImportJobLog]:
        self._job(workspace_id, job_id)
        return self.logs.list_for_job(workspace_id, job_id, status)

    def _job(self, workspace_id: UUID, job_id: UUID) -> ImportJob:
        job = self.jobs.get(workspace_id, job_id)
        if job is None:
            raise ImportServiceError("Import job not found")
        return job

    def _safe_filename(self, filename: str) -> str:
        return Path(filename).name.replace("/", "_").replace("\\", "_")


def map_row(raw_row: dict, column_mapping: dict[str, str]) -> dict:
    return {field: raw_row.get(column) for field, column in column_mapping.items() if column}


def parse_decimal(value) -> Decimal | None:
    return ExcelValueNormalizer().number(value)


def parse_datetime(value) -> datetime | None:
    return ExcelValueNormalizer().date(value)


def normalize_name(value: str) -> str:
    return sub(r"[^\wа-яА-ЯіїєґІЇЄҐ]+", "", value.lower())


def issue(row_number: int | None, severity: str, field: str | None, message: str, raw_value=None, normalized_value=None) -> ImportValidationIssue:
    return ImportValidationIssue(row_number=row_number, severity=severity, field=field, message=message, raw_value=raw_value, normalized_value=normalized_value)


def report_from_issues(rows: list[dict], issues: list[ImportValidationIssue]) -> ImportValidationReport:
    errors = [item.message for item in issues if item.severity == "ERROR"]
    warnings = [item.message for item in issues if item.severity == "WARNING"]
    return ImportValidationReport(is_valid=not errors, total_rows=len(rows), errors=errors, warnings=warnings, issues=issues)


def import_report(job_id: UUID, entity_type: str, sheet_name: str, rows: list[dict], validation: ImportValidationReport, extra_metrics: dict[str, int] | None = None) -> ImportReportResponse:
    issues_by_row: dict[int, list[ImportValidationIssue]] = defaultdict(list)
    for item in validation.issues:
        if item.row_number is not None:
            issues_by_row[item.row_number].append(item)
    error_rows = {row for row, issues in issues_by_row.items() if any(item.severity == "ERROR" for item in issues)}
    warning_rows = {row for row, issues in issues_by_row.items() if any(item.severity == "WARNING" for item in issues) and row not in error_rows}
    if entity_type == "product_catalog":
        duplicate_rows = {row for row, issues in issues_by_row.items() if any(item.field == "variant_sku" and "Duplicate" in item.message for item in issues)}
    else:
        duplicate_rows = {row for row, issues in issues_by_row.items() if any("Duplicate" in item.message for item in issues)}
    skipped_rows = duplicate_rows
    ready = max(len(rows) - len(error_rows) - len(skipped_rows), 0)
    return ImportReportResponse(
        job_id=job_id,
        entity_type=entity_type,
        sheet_name=sheet_name,
        total_rows=len(rows),
        valid_rows=len(rows) - len(error_rows),
        warning_rows=len(warning_rows),
        error_rows=len(error_rows),
        skipped_rows=len(skipped_rows),
        duplicate_rows=len(duplicate_rows),
        ready_to_import_rows=ready,
        estimated_entities_to_create=ready,
        products_detected=(extra_metrics or {}).get("products_detected", 0),
        variants_detected=(extra_metrics or {}).get("variants_detected", 0),
        inventory_rows_detected=(extra_metrics or {}).get("inventory_rows_detected", 0),
        images_detected=(extra_metrics or {}).get("images_detected", 0),
        duplicate_products=(extra_metrics or {}).get("duplicate_products", 0),
        duplicate_variants=(extra_metrics or {}).get("duplicate_variants", 0),
        sample_errors=[item for item in validation.issues if item.severity == "ERROR"][:5],
        sample_warnings=[item for item in validation.issues if item.severity == "WARNING"][:5],
    )


def safe_job_snapshot(job: ImportJob) -> dict:
    data = snapshot(job)
    data.pop("file_path", None)
    return data
